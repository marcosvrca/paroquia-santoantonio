import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


PRECO_BLOCO = 50.0


@dataclass
class RifaVendedorLinha:
    nome: str
    valor_rifas: float
    premiacao: float | None = None


@dataclass
class RifaPlanilha:
    doacoes: list[tuple[str, float]] = field(default_factory=list)
    despesas: list[tuple[str, float]] = field(default_factory=list)
    vendedores: list[RifaVendedorLinha] = field(default_factory=list)
    arrecadado_total: float | None = None


def _sheet_por_nome(wb, *palavras: str):
    for name in wb.sheetnames:
        norm = name.lower()
        if all(p in norm for p in palavras):
            return wb[name]
    return None


def _valor_celula(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip()
        if not txt or txt == "?":
            return None
        txt = txt.replace("R$", "").strip().replace(".", "").replace(",", ".")
        try:
            return float(txt)
        except ValueError:
            return None
    return None


def _parse_secao_linhas(ws, inicio_label: str, fim_label: str) -> list[tuple[str, float]]:
    itens: list[tuple[str, float]] = []
    capturando = False

    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if label is None:
            continue
        label_txt = str(label).strip()
        label_up = label_txt.upper()

        if inicio_label.upper() in label_up:
            capturando = True
            continue
        if not capturando:
            continue
        if fim_label.upper() in label_up:
            break

        valor = _valor_celula(ws.cell(r, 3).value)
        if valor is None or valor <= 0:
            continue
        itens.append((label_txt, round(valor, 2)))

    return itens


def _parse_prestacao_contas(ws) -> tuple[list[tuple[str, float]], list[tuple[str, float]], float | None]:
    doacoes = _parse_secao_linhas(ws, "DOA", "TOTAL (A)")
    despesas = _parse_secao_linhas(ws, "DESPESA", "TOTAL (B)")

    arrecadado = None
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if label and "VALOR ARRECADO" in str(label).upper():
            arrecadado = _valor_celula(ws.cell(r, 3).value)
            break

    return doacoes, despesas, arrecadado


def _parse_incentivo(ws) -> list[RifaVendedorLinha]:
    vendedores: list[RifaVendedorLinha] = []
    header_row = None

    for r in range(1, min(20, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, 5)]
        row_txt = " ".join(str(v or "") for v in vals).upper()
        if "NOME" in row_txt and "RIFA" in row_txt:
            header_row = r
            break

    if header_row is None:
        return vendedores

    for r in range(header_row + 1, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        rifas = _valor_celula(ws.cell(r, 2).value)
        premiacao = _valor_celula(ws.cell(r, 4).value)

        if nome is None and rifas is None:
            continue
        if nome is None:
            continue

        nome_txt = str(nome).strip()
        if not nome_txt or nome_txt.upper().startswith("TOTAL"):
            continue
        if rifas is None or rifas <= 0:
            continue

        vendedores.append(
            RifaVendedorLinha(
                nome=nome_txt,
                valor_rifas=round(rifas, 2),
                premiacao=round(premiacao, 2) if premiacao is not None else None,
            )
        )

    return vendedores


def parse_rifa_xlsx(path: str | Path) -> RifaPlanilha:
    wb = load_workbook(path, data_only=True)
    planilha = RifaPlanilha()

    ws_prest = _sheet_por_nome(wb, "presta") or (wb[wb.sheetnames[0]] if wb.sheetnames else None)
    if ws_prest is not None:
        planilha.doacoes, planilha.despesas, planilha.arrecadado_total = _parse_prestacao_contas(ws_prest)

    ws_inc = _sheet_por_nome(wb, "incentivo")
    if ws_inc is not None:
        planilha.vendedores = _parse_incentivo(ws_inc)

    if not planilha.doacoes and not planilha.despesas and not planilha.vendedores:
        raise ValueError("Não foi possível ler doações, despesas ou vendedores na planilha.")

    return planilha


def blocos_de_rifas(valor_rifas: float) -> float:
    return round(valor_rifas / PRECO_BLOCO, 2)
